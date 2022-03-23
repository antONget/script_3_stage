# подключаем библиотеки
import numpy as np
import pandas as pd
from openpyxl import Workbook
import csv
import pickle
from sklearn.metrics import confusion_matrix
from keras.models import Model
from keras.layers import Input, Dense
import os

# загружаем частотные срезы спектрограмм с сигналами phantom4
x_data256 = np.load('data/x_data256.npy')
# Загружаем разметку частотных срезов phantom4
y_data256 = np.load('data/y_data256.npy')
# загружаем частотные срезы фоновых спектрограмм
x_data256_big = np.load('data/x_data256_big.npy')

# инициализируем переменные
num = 25
x_train = x_data256[:95, num, :]


def create_dense_ae():
    # Размерность кодированного представления
    encoding_dim = 3
    # Энкодер
    # Входной плейсхолдер
    input_img = Input(shape=(x_train.shape[1]))
    encoded = Dense(150, activation='relu')(input_img)
    encoded = Dense(50, activation='relu')(encoded)
    encoded = Dense(encoding_dim, activation='relu')(encoded)
    # Декодер
    # Раскодированное другим полносвязным слоем изображение
    input_encoded = Input(shape=(encoding_dim,))
    decoded = Dense(150, activation='relu')(input_encoded)
    decoded = Dense(50, activation='relu')(decoded)
    decoded = Dense(x_train.shape[1], activation='relu')(decoded)
    # Модели, в конструктор первым аргументом передаются входные слои, а вторым выходные слои
    # Другие модели можно так же использовать как и слои
    encoder = Model(input_img, encoded, name="encoder")
    decoder = Model(input_encoded, decoded, name="decoder")
    autoencoder = Model(input_img, decoder(encoder(input_img)), name="autoencoder")
    return encoder, decoder, autoencoder


csv_np = np.zeros(shape=(256, 118))
count_train = 900
flag = 1  # обучаем(1) или используем готовую сеть(0)
if flag == 1:
    model = []
else:
    with open('model.data', 'rb') as filehandle:
        # сохраняем данные как двоичный поток
        model = pickle.load(filehandle)
for i in range(256):
    print(i)
    # iri = x_data256[y_data256[i, :] == 1, i, :]   # данные с сигналом
    # norm = x_data256[y_data256[i, :] == 0, i, :]  # данные с фоном
    x_train = x_data256_big[:count_train, i, :]   # обучающие данные
    if flag == 1:
        encoder, decoder, autoencoder = create_dense_ae()   # формируем сеть и обучаем ее
        autoencoder.compile(optimizer='adam', loss='binary_crossentropy')
        autoencoder.fit(x_train, x_train,
                        batch_size=10,
                        epochs=10, verbose=0)
        model.append(autoencoder)
    else:
        autoencoder = model[i]
    # рассчет ошибки восстановления фона
    pred_norm = autoencoder.predict(x_train)
    # pred_norm = x_train - pred_norm
    sum_err_norm = np.sum(pred_norm, axis=1)
    porog = np.mean(sum_err_norm)

    # рассчёт качества распознования
    csv_pred = autoencoder.predict(x_data256[:, i, :])
    csv_pred = x_data256[:, i, :] - csv_pred
    csv_pred = np.sum(csv_pred, axis=1)
    csv_np[i, csv_pred > porog] = 1

# преобразуем в int, для удобства работы
csv_np = csv_np.astype(int)
y_data256 = y_data256.astype(int)

if not os.path.exists('csv'):
    os.makedirs('csv')
if not os.path.exists('xlsx'):
    os.makedirs('xlsx')
# рассчет и сохранение матрицы ошибок предсказаных значений меток
matrix_error = y_data256 - csv_np
df_matrix_error = pd.DataFrame(matrix_error)
df_matrix_error.to_csv('csv/dif_matrix_error.csv')
wb = Workbook()
ws = wb.active
with open('csv/dif_matrix_error.csv', 'r') as f:
    for row in csv.reader(f):
        ws.append(row)
wb.save('xlsx/matrix_error.xlsx')

# вывод информации об ошибках
y_true = y_data256.reshape(-1, 1)
y_pred = csv_np.reshape(-1, 1)
tn, fp, fn, tp = confusion_matrix(y_true, y_pred).ravel()
print('Правильное обнаружение: ' + str(round((tp+tn) / len(y_pred)*100, 2))+'%')
print('Ложная тревога: ' + str(round(fp / len(y_pred)*100, 2))+'%')
print('Пропуск цели: ' + str(round(fn / len(y_pred)*100, 2))+'%')

# сохранение предсказанных значений меток
df_pred = pd.DataFrame(csv_np)
df_pred.to_csv('csv/df_pred.csv')
wb = Workbook()
ws = wb.active
with open('csv/df_pred.csv', 'r') as f:
    for row in csv.reader(f):
        ws.append(row)
wb.save('xlsx/df_pred.xlsx')

# сохранение истинных значений меток
df_y_true = pd.DataFrame(y_data256)
df_y_true.to_csv('csv/df_y_true.csv')
wb = Workbook()
ws = wb.active
with open('csv/df_y_true.csv', 'r') as f:
    for row in csv.reader(f):
        ws.append(row)
wb.save('xlsx/df_y_true.xlsx')

# преобразование массива предсказанных меток в соответствии с правилом 5из5
# если в 5 последовательных спектрограммах на одном частотном срезе предсказан сигнал,
# то  решение принято верно
solution = np.zeros(shape=(256, 118))
for row in range(csv_np.shape[0]):
    for num in range(csv_np.shape[1] - 5):
        if np.sum(csv_np[row, num:num + 5]) == 5.:
            solution[row, num:num + 5] = 1

# вывод информации об ошибках
y_true = y_data256.reshape(-1, 1)
y_pred = solution.reshape(-1, 1)
tn, fp, fn, tp = confusion_matrix(y_true, y_pred).ravel()
print('--- Алгоритм 5 из 5 ---')
print('Правильное обнаружение: ' + str(round((tp+tn) / len(y_pred)*100, 2))+'%')
print('Ложная тревога: ' + str(round(fp / len(y_pred)*100, 2))+'%')
print('Пропуск цели: ' + str(round(fn / len(y_pred)*100, 2))+'%')

# if not os.path.exists('xlsx'):
#     os.makedirs('xlsx')
# # сохранение преобразованного массива предсказанных меток
# df_solution = pd.DataFrame(solution)
# df_solution.to_csv('csv/df_solution.csv')
# wb = Workbook()
# ws = wb.active
# with open('csv/df_solution.csv', 'r') as f:
#     for row in csv.reader(f):
#         ws.append(row)
# wb.save('xlsx/df_solution.xlsx')
#
# # сохранение массива матрицы ошибок
# matrix_errorS = y_data256 - solution
# df_matrix_errorS = pd.DataFrame(matrix_errorS)
# df_matrix_errorS.to_csv('csv/dif_matrix_error_solution.csv')
# wb = Workbook()
# ws = wb.active
# with open('csv/dif_matrix_error_solution.csv', 'r') as f:
#     for row in csv.reader(f):
#         ws.append(row)
# wb.save('xlsx/matrix_error_solution.xlsx')
#
# # создаем exel таблицу для визуального представления
# folder = r'xlsx'
# path = os.path.join(folder, 'excels.xlsx')
# writer = pd.ExcelWriter(path, engine='openpyxl')
# for i in os.listdir(folder)[:-1]:
#     pd.read_excel(os.path.join(folder, i)).to_excel(writer, i[:-5], index=False)
#     writer.save()
# if flag:
#     with open('model.data', 'wb') as filehandle:
#         # сохраняем данные как двоичный поток
#         pickle.dump(model, filehandle)
# print('всё')
